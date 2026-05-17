#!/usr/bin/env python3
"""
Extract AHSP Cipta Karya data from SE Bina Konstruksi No. 68/2024 Excel file.

Outputs one JSON file per sheet into:
  packages/cipta-karya-2024/data/ahsp/divisi-X/<sheet-slug>.json

Each JSON file is an array of AhspItem objects (fixed_coefficient type).

Layout detection per sheet:
  - Layout A : col[4]="Kode", col[6]=Koefisien, col[7]=HargaSatuan, col[8 or 9]=Jumlah
  - Layout B : col[4]=Satuan,  col[5]=Koefisien, col[6]=HargaSatuan, col[7]=Jumlah
  - Layout C : Lansekap (HSD calculation sheet) - skip
"""

import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

XLSX = "/root/.claude/uploads/64129fc8-b021-484f-b751-749ea0ddfb00/ed24cada-AHSP_CIPTA_KARYA_SE_BINA_KONSTRUKSI_NO_68_TAHUN_2024.xlsx"
OUT_ROOT = Path(__file__).parent.parent / "packages" / "cipta-karya-2024" / "data" / "ahsp"
SUMBER = "SE Bina Konstruksi No. 68/SE/Dk/2024"

# Mapping sheet names → (divisi, slug)
SHEET_MAP = {
    "Persiapan":                      (1, "divisi-1", "persiapan"),
    "Galian Tanah":                   (1, "divisi-1", "galian-tanah"),
    "Timbunan Pemadatan":             (2, "divisi-2", "timbunan-pemadatan"),
    "Angkut Material":                (2, "divisi-2", "angkut-material"),
    "Geotekstil & Geomembran ":       (2, "divisi-2", "geotekstil-geomembran"),
    "Pembongkaran":                   (3, "divisi-3", "pembongkaran"),
    "Rangka Atap":                    (3, "divisi-3", "rangka-atap"),
    "Beton":                          (4, "divisi-4", "beton"),
    "Pondasi":                        (4, "divisi-4", "pondasi"),
    "BAJA":                           (4, "divisi-4", "baja"),
    "Beton Pracetak":                 (4, "divisi-4", "beton-pracetak"),
    "Beton Prategang ":               (4, "divisi-4", "beton-prategang"),
    "Struktur Kayu":                  (4, "divisi-4", "struktur-kayu"),
    "Dinding Penahan Tanah ":         (4, "divisi-4", "dinding-penahan-tanah"),
    "Penutup Atap":                   (5, "divisi-5", "penutup-atap"),
    "Plafon":                         (5, "divisi-5", "plafon"),
    "Pasangan Dinding":               (5, "divisi-5", "pasangan-dinding"),
    "Plesteran Dan Acian":            (5, "divisi-5", "plesteran-acian"),
    "Pengecatan dan Pelituran":       (5, "divisi-5", "pengecatan-pelituran"),
    "Penutup Lantai dan Dinding":     (5, "divisi-5", "penutup-lantai-dinding"),
    "Pintu dan Jendela":              (5, "divisi-5", "pintu-jendela"),
    "Kaca":                           (5, "divisi-5", "kaca"),
    "Besi dan Aluminium":             (5, "divisi-5", "besi-aluminium"),
    "Kayu":                           (5, "divisi-5", "kayu"),
    "Ornamen":                        (5, "divisi-5", "ornamen"),
    "Signage":                        (5, "divisi-5", "signage"),
    "Sanitair":                       (6, "divisi-6", "sanitair"),
    "Lansekap":                       (6, "divisi-6", "lansekap"),  # special layout, may skip
    "Jaringan Listrik":               (7, "divisi-7", "jaringan-listrik"),
    "Perpipaan dan proteksi kebakar ": (8, "divisi-8", "perpipaan-proteksi-kebakaran"),
    "Sistem Air Minum":               (8, "divisi-8", "sistem-air-minum"),
    "Sistem Air Limbah":              (8, "divisi-8", "sistem-air-limbah"),
    "Bak Kontrol":                    (8, "divisi-8", "bak-kontrol"),
    "Perpipaan dalam Gedung ":        (8, "divisi-8", "perpipaan-dalam-gedung"),
    "Sistem Air Hujan":               (8, "divisi-8", "sistem-air-hujan"),
    "Jalan Pada Pemukiman":           (9, "divisi-9", "jalan-pemukiman"),
    "Drainase":                       (9, "divisi-9", "drainase"),
    "Jaringan Pipa Luar Gedung":      (9, "divisi-9", "jaringan-pipa-luar-gedung"),
    "Strutur Risha ":                 (10, "divisi-10", "struktur-risha"),
}

SKIP_SHEETS = {"Daftar Harga Satuan Pekerjaan", "Upah Bahan", "Lansekap"}  # Lansekap is an HSD calc sheet

STANDARD_MARGIN = {
    "overhead_pct": {"label": "Biaya Umum (Overhead)", "min": 0, "max": 15, "default": 10},
    "profit_pct": {"label": "Keuntungan (Profit)", "min": 0, "max": 15, "default": 5},
    "constraint": {"rule": "10 <= overhead_pct + profit_pct <= 15"},
}

# L-code lookup
TK_CODES = {
    "Pekerja": "L.01",
    "Tukang batu": "L.02", "Tukang kayu": "L.02", "Tukang besi": "L.02",
    "Tukang cat": "L.02", "Tukang pipa": "L.02", "Tukang las": "L.02",
    "Tukang listrik": "L.02", "Tukang alumunium": "L.02", "Tukang kaca": "L.02",
    "Tukang erection": "L.02", "Tukang tanam": "L.02", "Tukang pelitur": "L.02",
    "Tukang pemelihara taman": "L.02",
    "Kepala tukang": "L.03",
    "Mandor": "L.04",
    "Juru ukur": "L.05",
    "Pembantu juru ukur": "L.06",
    "Mekanik alat berat": "L.07",
    "Operator alat berat": "L.08",
    "Pembantu operator alat berat": "L.09",
    "Sopir": "L.10",
    "Kenek": "L.11",
}


def clean_str(v):
    if v is None:
        return None
    return str(v).strip()


def is_kode_ahsp(v):
    """Return True if value looks like an AHSP code (e.g. '1.2.1.1.1', '2.2')."""
    if v is None:
        return False
    s = str(v).strip()
    return bool(re.match(r'^\d+\.\d+', s)) and not s.startswith('0')


def detect_layout(ws):
    """
    Scan first few hundred rows to find a header row and determine layout.
    Returns 'A', 'B', or 'C'.
    Also returns the column index for:
       no_col, uraian_col, kode_col (or None), sat_col, koef_col, harga_col, jumlah_col
    """
    for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
        vals = [clean_str(v) for v in row]
        # Find a header row containing "Koefisien"
        if "Koefisien" in vals or "koefisien" in vals.lower() if hasattr(vals, 'lower') else False:
            pass
        koef_col = None
        for i, v in enumerate(vals):
            if v and v.lower() == "koefisien":
                koef_col = i
                break
        if koef_col is None:
            continue

        # We found the header row. Determine layout by col[4]
        col4 = vals[4] if len(vals) > 4 else None

        if col4 and "Kode" in col4:
            # Layout A: No=2, Uraian=3, Kode=4, Sat=5, Koef=6, Harga=7, Jumlah=?
            # Find "Jumlah" column
            jumlah_col = None
            for i, v in enumerate(vals):
                if v and "Jumlah" in v and "HARGA" not in v.upper():
                    jumlah_col = i
                    break
            if jumlah_col is None:
                # scan for column containing "Jumlah Harga"
                for i, v in enumerate(vals):
                    if v and "Jumlah" in v:
                        jumlah_col = i
                        break
            return {
                "layout": "A",
                "no_col": 2, "uraian_col": 3, "kode_col": 4,
                "sat_col": 5, "koef_col": 6, "harga_col": 7,
                "jumlah_col": jumlah_col or 8,
            }
        elif col4 and ("Sat" in col4 or "sat" in col4.lower()):
            # Layout B: No=2, Uraian=3, Sat=4, Koef=5, Harga=6, Jumlah=7
            return {
                "layout": "B",
                "no_col": 2, "uraian_col": 3, "kode_col": None,
                "sat_col": 4, "koef_col": 5, "harga_col": 6,
                "jumlah_col": 7,
            }
        else:
            # Fallback: try to detect from koef_col position
            if koef_col == 5:
                return {
                    "layout": "B",
                    "no_col": 2, "uraian_col": 3, "kode_col": None,
                    "sat_col": 4, "koef_col": 5, "harga_col": 6,
                    "jumlah_col": 7,
                }
            elif koef_col == 6:
                return {
                    "layout": "A",
                    "no_col": 2, "uraian_col": 3, "kode_col": 4,
                    "sat_col": 5, "koef_col": 6, "harga_col": 7,
                    "jumlah_col": 8,
                }
    # Couldn't detect - mark as C (skip)
    return {"layout": "C"}


def get_section_type(val):
    """Return 'tk', 'bahan', 'peralatan', 'end', or None."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ("A", "TENAGA KERJA"):
        return "tk"
    if s in ("B", "BAHAN", "MATERIAL"):
        return "bahan"
    if s in ("C", "PERALATAN", "ALAT"):
        return "peralatan"
    if s in ("D", "E", "F") or "JUMLAH" in s or "BIAYA UMUM" in s or "HARGA SATUAN PEKERJAAN" in s:
        return "end"
    return None


def parse_sheet(ws, sheet_name, layout_info, divisi, divisi_dir):
    """Parse a single AHSP sheet and return list of AhspItem dicts."""
    if layout_info["layout"] == "C":
        print(f"  [SKIP] {sheet_name} — Layout C (HSD calculation)")
        return []

    lyt = layout_info
    items = []
    current_item = None
    current_section = None  # 'tk', 'bahan', 'peralatan'

    rows = list(ws.iter_rows(min_row=1, values_only=True))

    i = 0
    while i < len(rows):
        row = rows[i]
        # Get value at uraian_col (col 3)
        col2 = row[2] if len(row) > 2 else None
        col3 = row[3] if len(row) > 3 else None

        col2_str = clean_str(col2)
        col3_str = clean_str(col3)

        # Detect new AHSP item: col[2] is a code like "1.2.1.1.1"
        if is_kode_ahsp(col2_str):
            # Save previous item
            if current_item is not None:
                items.append(current_item)

            # Extract the final price from the end of this row or next few rows
            # The final price is usually at a specific column in the same row
            # Look for the rightmost numeric column
            final_price = None
            for colidx in range(len(row) - 1, 4, -1):
                v = row[colidx]
                if isinstance(v, (int, float)) and v > 0:
                    final_price = round(float(v))
                    break

            # Determine divisi and sub_divisi from code
            code = col2_str.strip()
            parts = code.split(".")
            item_divisi = int(parts[0]) if parts else divisi

            # sub_divisi is first two parts: e.g. "1.2"
            sub_divisi = f"{parts[0]}.{parts[1]}" if len(parts) >= 2 else code

            # Determine jenis_pekerjaan from name
            nama = col3_str or ""
            jp = "manual"
            if "semi mekanis" in nama.lower() or "semi-mekanis" in nama.lower():
                jp = "semi-mekanis"
            elif "mekanis" in nama.lower() and "semi" not in nama.lower():
                jp = "mekanis"

            current_item = {
                "kode_ahsp": code,
                "nama": nama,
                "bidang": "cipta-karya",
                "divisi": item_divisi,
                "sub_divisi": sub_divisi,
                "satuan_bayar": "",  # will be filled below
                "jenis_kalkulasi": "fixed_coefficient",
                "jenis_pekerjaan": jp,
                "is_lump_sum": False,
                "tenaga_kerja": [],
                "bahan": [],
                "peralatan": [],
                "harga_satuan_pekerjaan_ref": final_price,
                "margin": STANDARD_MARGIN,
                "provenance": {
                    "sumber_regulasi": SUMBER,
                    "halaman": sheet_name,
                    "diverifikasi_oleh": None,
                    "tanggal_verifikasi": None,
                },
            }
            current_section = None
            i += 1
            continue

        # Detect section headers: A, B, C in col[2]
        if current_item is not None and col2_str:
            st = get_section_type(col2_str)
            if st in ("tk", "bahan", "peralatan"):
                current_section = st
                i += 1
                continue
            elif st == "end":
                # Look for satuan_bayar in the "F" row
                if col2_str.upper() == "F" and col3_str and "Harga Satuan" in col3_str:
                    # Check if satuan is embedded in the description
                    pass
                i += 1
                continue

        # Parse component rows
        if current_item is not None and current_section is not None:
            # Try to extract component data
            # For Layout A: row = [..., No, Uraian, Kode, Sat, Koef, Harga, Jumlah, ...]
            # For Layout B: row = [..., No, Uraian, Sat, Koef, Harga, Jumlah, ...]
            try:
                uraian = clean_str(row[lyt["uraian_col"]])
                if not uraian or uraian.upper() in ("JUMLAH HARGA TENAGA KERJA", "JUMLAH HARGA BAHAN",
                                                      "JUMLAH HARGA ALAT", "JUMLAH HARGA PERALATAN",
                                                      "CATATAN:"):
                    i += 1
                    continue

                # Skip header rows within sections
                if uraian.lower() in ("uraian", "no", "a", "b", "c", "d", "e", "f"):
                    i += 1
                    continue

                # Check if this is a sub-section header (all caps, no number)
                if uraian.isupper() and not any(isinstance(row[j], (int, float)) for j in range(lyt["koef_col"], min(lyt["koef_col"] + 3, len(row)))):
                    i += 1
                    continue

                if lyt["layout"] == "A":
                    kode_ref = clean_str(row[lyt["kode_col"]]) if lyt["kode_col"] < len(row) else None
                    satuan = clean_str(row[lyt["sat_col"]]) if lyt["sat_col"] < len(row) else None
                    koef = row[lyt["koef_col"]] if lyt["koef_col"] < len(row) else None
                    harga = row[lyt["harga_col"]] if lyt["harga_col"] < len(row) else None
                else:  # Layout B
                    kode_ref = None
                    satuan = clean_str(row[lyt["sat_col"]]) if lyt["sat_col"] < len(row) else None
                    koef = row[lyt["koef_col"]] if lyt["koef_col"] < len(row) else None
                    harga = row[lyt["harga_col"]] if lyt["harga_col"] < len(row) else None

                # Validate
                if koef is None or not isinstance(koef, (int, float)) or koef <= 0:
                    i += 1
                    continue
                if harga is None or not isinstance(harga, (int, float)):
                    i += 1
                    continue

                koef = float(koef)
                harga = float(harga)

                # Extract satuan_bayar from the AHSP item name if not yet set
                if not current_item["satuan_bayar"]:
                    # Try to infer from the item name: "1 m3 ...", "per m2 ..."
                    name_lower = current_item["nama"].lower()
                    sat_match = re.search(r'\b(m3|m2|m\'|m1|m|kg|buah|btg|set|ls|ha|tgl|hari|jam|unit|lbr|bh|bj|bk|bl|titik|sambungan|buah)\b', name_lower)
                    if sat_match:
                        current_item["satuan_bayar"] = sat_match.group(1)

                if current_section == "tk":
                    entry = {
                        "nama": uraian,
                        "satuan": satuan or "OH",
                        "koefisien": round(koef, 6),
                        "harga_satuan_ref": round(harga),
                    }
                    # Try to resolve ref code
                    if kode_ref and re.match(r'^L\.', kode_ref):
                        entry["ref"] = kode_ref
                    else:
                        # Lookup from name
                        ref = TK_CODES.get(uraian)
                        if ref:
                            entry["ref"] = ref
                        else:
                            entry["ref"] = None
                    current_item["tenaga_kerja"].append(entry)

                elif current_section == "bahan":
                    entry = {
                        "nama": uraian,
                        "satuan": satuan or "",
                        "koefisien": round(koef, 6),
                        "harga_satuan_ref": round(harga),
                    }
                    if kode_ref and re.match(r'^M\.', kode_ref):
                        entry["ref"] = kode_ref
                    current_item["bahan"].append(entry)

                elif current_section == "peralatan":
                    entry = {
                        "nama": uraian,
                        "satuan": satuan or "",
                        "koefisien": round(koef, 6),
                        "harga_satuan_ref": round(harga),
                    }
                    if kode_ref and re.match(r'^E\.', kode_ref):
                        entry["ref"] = kode_ref
                    current_item["peralatan"].append(entry)

            except (IndexError, TypeError):
                pass

        i += 1

    # Append last item
    if current_item is not None:
        items.append(current_item)

    # Post-process: remove section headers (no price, no components)
    items = [
        item for item in items
        if item.get("harga_satuan_pekerjaan_ref") or item["tenaga_kerja"] or item["bahan"] or item["peralatan"]
    ]

    # Fill missing satuan_bayar
    for item in items:
        if not item["satuan_bayar"]:
            item["satuan_bayar"] = "unit"

    return items


def main():
    print(f"Loading {XLSX}")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    total_items = 0
    total_files = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        if sheet_name not in SHEET_MAP:
            print(f"  [WARN] Unknown sheet: {repr(sheet_name)} — skipping")
            continue

        divisi, divisi_dir_name, slug = SHEET_MAP[sheet_name]
        ws = wb[sheet_name]

        layout_info = detect_layout(ws)
        print(f"Processing [{sheet_name}] → Layout={layout_info['layout']}, divisi={divisi}")

        if layout_info["layout"] == "C":
            print(f"  [SKIP] Lansekap-style sheet")
            continue

        items = parse_sheet(ws, sheet_name, layout_info, divisi, divisi_dir_name)

        if not items:
            print(f"  [WARN] No items extracted from {sheet_name}")
            continue

        # Write output
        out_dir = OUT_ROOT / divisi_dir_name
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / f"{slug}.json"

        with open(out_file, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)

        total_items += len(items)
        total_files += 1
        print(f"  → {len(items)} items → {out_file.relative_to(Path(__file__).parent.parent)}")

    print(f"\nDone: {total_items} items across {total_files} files.")


if __name__ == "__main__":
    main()
